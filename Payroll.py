import sys
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QVBoxLayout, QScrollArea, QHBoxLayout,
    QFrame, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QGridLayout,
    QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon, QFont
import mainsyslogics
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


class PayrollDetailWindow(QWidget):
    """Window to show detailed payroll for a specific week"""

    def __init__(self, week_start, companyName):
        super().__init__()
        self.setWindowTitle(f"Payroll Details - {week_start}")
        self.resize(1000, 600)

        self.week_start = week_start
        self.companyName = companyName
        self.payroll_data = []

        layout = QVBoxLayout(self)

        # Header
        header = QLabel(f"<h2>Payroll for Week: {week_start}</h2>")
        header.setStyleSheet("color: #2c3e50;")
        layout.addWidget(header, alignment=Qt.AlignmentFlag.AlignCenter)

        # Table
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #345e3d;
                color: white;
                gridline-color: #7f8c8d;
                border: 1px solid #7f8c8d;
            }
            QTableWidget::item {
                color: white;
            }
            QHeaderView::section {
                background-color: #1abc9c;
                color: white;
                padding: 8px;
                border: 1px solid #16a085;
                font-weight: bold;
            }
        """)
        layout.addWidget(self.table)

        # Button layout
        button_layout = QHBoxLayout()

        # Export to Excel button
        export_btn = QPushButton("📊 Export to Excel")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #229954; }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(export_btn)

        button_layout.addStretch()

        # Close button
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #c0392b; }
        """)
        close_btn.clicked.connect(self.close)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)

        self.load_payroll_data()

    def load_payroll_data(self):
        db = mainsyslogics.Connectdb()
        try:
            self.payroll_data = db.get_payroll_details(self.week_start, self.companyName)
        except Exception as e:
            print(f"❌ Error loading payroll: {e}")
            self.payroll_data = []

        headers = ["Employee Name", "Total Hours", "Hourly Rate", "Gross Pay", "Deductions", "Net Pay"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.payroll_data))

        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)

        total_gross = 0
        total_deductions = 0
        total_net = 0

        for row_idx, data in enumerate(self.payroll_data):
            self.table.setItem(row_idx, 0, QTableWidgetItem(data['employee_name']))
            self.table.setItem(row_idx, 1, QTableWidgetItem(f"{data['total_hours']:.2f}"))
            self.table.setItem(row_idx, 2, QTableWidgetItem(f"₱{data['daily_rate']:.2f}"))
            self.table.setItem(row_idx, 3, QTableWidgetItem(f"₱{data['gross_pay']:.2f}"))
            self.table.setItem(row_idx, 4, QTableWidgetItem(f"₱{data['deductions']:.2f}"))
            self.table.setItem(row_idx, 5, QTableWidgetItem(f"₱{data['net_pay']:.2f}"))

            total_gross += data['gross_pay']
            total_deductions += data['deductions']
            total_net += data['net_pay']

        # Add totals row
        total_row = self.table.rowCount()
        self.table.insertRow(total_row)
        totals = ["TOTAL", "", "", f"₱{total_gross:.2f}", f"₱{total_deductions:.2f}", f"₱{total_net:.2f}"]
        for col, text in enumerate(totals):
            item = QTableWidgetItem(text)
            item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            item.setBackground(Qt.GlobalColor.lightGray)
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight if col != 0 else Qt.AlignmentFlag.AlignLeft)
            self.table.setItem(total_row, col, item)

        # Store totals for export
        self.total_gross = total_gross
        self.total_deductions = total_deductions
        self.total_net = total_net

    def export_to_excel(self):
        """Export payroll data to Excel file"""
        try:
            # Ask user where to save
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Payroll Report",
                f"Payroll_{self.companyName}_{self.week_start}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Payroll Report"

            # Header styling
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Company header
            ws.merge_cells('A1:F1')
            ws['A1'] = f"{self.companyName} - Payroll Report"
            ws['A1'].font = Font(bold=True, size=16, color="2C3E50")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Week period
            ws.merge_cells('A2:F2')
            ws['A2'] = f"Week: {self.week_start}"
            ws['A2'].font = Font(bold=True, size=12, color="34495E")
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

            # Date generated
            ws.merge_cells('A3:F3')
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(size=10, color="7F8C8D")
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

            # Column headers
            headers = ["Employee Name", "Total Hours", "Hourly Rate", "Gross Pay", "Deductions", "Net Pay"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Data rows
            row_num = 6
            for data in self.payroll_data:
                ws.cell(row=row_num, column=1, value=data['employee_name']).border = border
                ws.cell(row=row_num, column=2, value=data['total_hours']).border = border
                ws.cell(row=row_num, column=3, value=data['daily_rate']).border = border
                ws.cell(row=row_num, column=4, value=data['gross_pay']).border = border
                ws.cell(row=row_num, column=5, value=data['deductions']).border = border
                ws.cell(row=row_num, column=6, value=data['net_pay']).border = border

                # Format currency columns
                for col in [3, 4, 5, 6]:
                    ws.cell(row=row_num, column=col).number_format = '₱#,##0.00'
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='right')

                # Format hours column
                ws.cell(row=row_num, column=2).number_format = '0.00'
                ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')

                row_num += 1

            # Totals row
            total_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            total_font = Font(bold=True, size=11)

            ws.cell(row=row_num, column=1, value="TOTAL").font = total_font
            ws.cell(row=row_num, column=1).fill = total_fill
            ws.cell(row=row_num, column=1).border = border

            ws.cell(row=row_num, column=2, value="").fill = total_fill
            ws.cell(row=row_num, column=2).border = border

            ws.cell(row=row_num, column=3, value="").fill = total_fill
            ws.cell(row=row_num, column=3).border = border

            ws.cell(row=row_num, column=4, value=self.total_gross).font = total_font
            ws.cell(row=row_num, column=4).fill = total_fill
            ws.cell(row=row_num, column=4).number_format = '₱#,##0.00'
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=4).border = border

            ws.cell(row=row_num, column=5, value=self.total_deductions).font = total_font
            ws.cell(row=row_num, column=5).fill = total_fill
            ws.cell(row=row_num, column=5).number_format = '₱#,##0.00'
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=5).border = border

            ws.cell(row=row_num, column=6, value=self.total_net).font = total_font
            ws.cell(row=row_num, column=6).fill = total_fill
            ws.cell(row=row_num, column=6).number_format = '₱#,##0.00'
            ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=6).border = border

            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15

            # Save workbook
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Successful",
                f"Payroll report exported successfully to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export payroll report:\n{str(e)}"
            )
            print(f"❌ Error exporting to Excel: {e}")


# ---------------- Payroll Card (Matching Attendance Style) ----------------
class PayrollCard(QFrame):
    def __init__(self, week_start, summary, total_net_pay, companyName, parent=None):
        super().__init__(parent)
        self.week_start = week_start
        self.companyName = companyName
        self.parent_window = parent
        self.setFixedHeight(80)

        # Green color matching submitted attendance
        bg_color = "rgb(31, 158, 37)"

        self.setStyleSheet(f"""
            QFrame {{
                border: 2px solid #2c3e50;
                border-radius: 25px;
                background-color: {bg_color};
                color: white;
            }}
        """)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(20, 10, 20, 10)

        week_label = QLabel(f"<b>{week_start}</b>")
        week_label.setStyleSheet("font-size: 16px;")

        status_label = QLabel(f"[PAID]")
        status_label.setStyleSheet("font-size: 12px;")

        summary_label = QLabel(summary)
        summary_label.setStyleSheet("font-size: 13px;")

        net_pay_label = QLabel(f"<b>Net: ₱{total_net_pay:,.2f}</b>")
        net_pay_label.setStyleSheet("font-size: 14px;")

        layout.addWidget(week_label)
        layout.addWidget(status_label)
        layout.addStretch()
        layout.addWidget(summary_label)
        layout.addWidget(net_pay_label)

        self.button = QPushButton("", self)
        self.button.setFlat(True)
        self.button.setStyleSheet("background: transparent;")
        self.button.clicked.connect(self.open_detail)

    def resizeEvent(self, event):
        self.button.setGeometry(0, 0, self.width(), self.height())
        super().resizeEvent(event)

    def open_detail(self):
        self.detail_window = PayrollDetailWindow(
            self.week_start,
            self.companyName
        )
        self.detail_window.show()


class PayrollWindow(QWidget):
    """Main payroll summary window"""

    def __init__(self, role, companyName):
        super().__init__()
        self.setWindowTitle('LooTech - Payroll')
        self.setWindowIcon(QIcon('stylesAndPic/lootechIcon.png'))
        self.resize(900, 600)

        self.__role = role
        self.__companyName = companyName

        layout = QGridLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.body = QWidget()
        self.body_layout = QGridLayout(self.body)
        self.body.setObjectName('body')
        self.body_layout.setContentsMargins(10, 0, 10, 10)
        self.body_layout.setSpacing(10)

        label = QLabel("<h1>Payroll</h1>")
        label.setProperty('class', 'label')
        self.body_layout.addWidget(label, 0, 0, alignment=Qt.AlignmentFlag.AlignTop)

        layout.addWidget(self.body, 1, 0, 4, 1, alignment=Qt.AlignmentFlag.AlignTop)

        self.load_payroll_records()

    def load_payroll_records(self):
        # Clear existing cards
        for i in reversed(range(self.body_layout.count())):
            if i > 0:
                widget = self.body_layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()

        db = mainsyslogics.Connectdb()
        try:
            summaries = db.get_payroll_summary(self.__companyName)
        except Exception as e:
            print(f"❌ Error loading payroll: {e}")
            summaries = []

        if not summaries:
            lbl = QLabel("<i>No payroll records yet.<br>Submit attendance to generate payroll.</i>")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("color: #7f8c8d; padding: 40px; font-size: 14px;")
            self.body_layout.addWidget(lbl, 1, 0, 1, 5)
            return

        row = 1
        for summary in summaries:
            summary_text = f"Gross: ₱{summary['total_gross_pay']:,.2f} | Deductions: ₱{summary['total_deductions']:,.2f}"
            card = PayrollCard(
                summary['week_start'],
                summary_text,
                summary['total_net_pay'],
                self.__companyName,
                parent=self
            )
            self.body_layout.addWidget(card, row, 0, 1, 5)
            row += 1